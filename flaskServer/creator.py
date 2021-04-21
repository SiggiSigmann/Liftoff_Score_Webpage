import dbconnector.dbconnector as dbcon
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import sys
import re
import io


class Creator():
	user_list = []
	drone_list = []

	user_id = {}
	drone_id = {}

	def __init__(self, db):
		self.db = db
		for user in db.get_users():
			self.user_list.append(user)
		
		for drone in db.get_drones():
			self.drone_list.append([drone[0],drone[1]])

	def createFile(self):
		wb = Workbook()

		maps_track = self.db.get_tracks()
		best = self.db.get_best_reslt_per_user()

		ws = wb.active
		for drone in self.drone_list:
	
			ws.title = drone[1]


			ws.cell(row=1, column=1, value = "MAP")
			ws.cell(row=1, column=2, value = "Track")
			ws.cell(row=1, column=3, value = "Race")
			ws.cell(row=1, column=4, value = "Level")

			col = 5
			for user in self.user_list:
				ws.cell(row=1, column=col, value = user[1])
				print(user[2][1:], file=sys.stderr)
				ws.cell(row=1, column=col).fill = PatternFill(start_color=user[2][1:], end_color=user[2][1:], fill_type = "solid")
				
				col += 1

			row = 2
			for maps in maps_track["maps"]:
				for track in maps["tracks"]:
					ws.cell(row=row, column=1, value = maps["mapname"])
					ws.cell(row=row, column=2, value = track["trackname"])
					ws.cell(row=row, column=3, value = track["trackname"])
					ws.cell(row=row, column=4, value = track["hardness"])

					col = 5
					for user in self.user_list:
						if drone[0] in best.keys():
							print("drone",file=sys.stderr)
							if maps["mapid"] in best[drone[0]].keys():
								print("map",file=sys.stderr)
								if track["trackid"] in best[drone[0]][maps["mapid"]].keys():
									print("track",file=sys.stderr)
									if user[0] in best[drone[0]][maps["mapid"]][track["trackid"]].keys():
										print("user",file=sys.stderr)
										ws.cell(row=row, column=col, value = best[drone[0]][maps["mapid"]][track["trackid"]][user[0]])
						
						col += 1


					row += 1

			if self.drone_list[-1] != drone:
				ws = wb.create_sheet()
		return wb