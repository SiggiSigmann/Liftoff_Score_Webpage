import dbconnector.dbconnector as dbcon
from openpyxl import load_workbook
import sys
import re


class ExcelLoader():
	user_list = []
	drone_list = []

	excluded_headder=["Map","Track","Race","Level",None]

	user_id = {}
	drone_id = {}

	def __init__(self, db):
		self.db = db
		for user in db.get_users():
			self.user_list.append(user[1])
		
		for drone in db.get_drones():
			self.drone_list.append(drone[1])

	def loadFile(self, filename):
		workbook = load_workbook(filename=filename)

		#itterate over pages ( pagename = dronename)
		for dronename in workbook.sheetnames:
			#check drone
			success = self._make_shure_drone_exists(dronename)
			if success != 0:
				return success

			#get sheet
			sheet = workbook[dronename]
			firstrow = sheet[1]

			username_idx = {}
			index = 0
			user_count = 0

			#get header in file
			for cell in firstrow:
				#check if header is a user
				if not cell.value in self.excluded_headder:
					user = cell.value.rstrip()
					color = "#"+cell.fill.start_color.index[2:].rstrip()
					success = self._make_shure_user_exists(user, color)
					if success != 0:
						return success

					#stres which row correleates to which user
					username_idx[index] = user
					user_count += 1
				index += 1

			#used to check if more rows contains values
			max_idx = user_count+4
			
			#load id for user and drone form db
			self._load_user_id()
			self._load_drone_id()
			
			skip_first = 0
			for row in sheet:
				if skip_first == 0:
					skip_first = 1
					continue

				#check if row contains values
				if row[0].value != None:
					map_name = row[0].value.rstrip()
					track_name = row[1].value.rstrip()

					#itterate over users value
					index = 4
					for user in row[4:]:
						#check if a usere exists for this value
						if index < max_idx:
							#check if entry exists
							if user.value != None:
								time = user.value.rstrip()
								username = username_idx[index]

								

								#check value from file
								matched = re.match("[0-9]{2}:[0-9]{2}:[0-9]{3}", time)
								if not bool(matched):
									print(time, file = sys.stderr)
									print(username + " "+dronename+" "+map_name+" "+track_name, file = sys.stderr)
									return -5

								#get map and track id from db 
								ids = self.db.get_map_track_id(map_name, track_name)
								if len(ids) < 1:
									return -6


								self.db.add_new_result(ids[0], ids[1], str(self.user_id[username]), self.drone_id[dronename], time)
						index += 1
		return 0

	def _make_shure_user_exists(self, username, color):
		if not username in self.user_list:
			matched = re.match("[a-zA-z0-9 ]{3,25}", username)
			if not bool(matched):
				return -4
			print("user not exis: " + username + " ("+color+")", file=sys.stderr)
			self.db.add_new_user(username,color)
			self.user_list.append(username)
		return 0

	def _make_shure_drone_exists(self, dronename):
		if not dronename in self.drone_list:
			matched = re.match("[a-zA-z0-9 ]{3,25}", dronename)
			if not bool(matched):	
				return -3
			print("drone not exis: " + dronename, file=sys.stderr)
			self.db.add_new_drone(dronename)
			self.drone_list.append(dronename)
		return 0

	def _load_user_id(self):
		users =  self.db.get_users()
		for user in users:
			self.user_id[user[1]] = user[0]

	def _load_drone_id(self):
		drones = self.db.get_drones()
		for drone in drones:
			self.drone_id[drone[1]] = drone[0]
		
			


			

		